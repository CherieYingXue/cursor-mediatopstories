import csv
import datetime as dt
import json
import os
import sqlite3
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urljoin, urlparse

import feedparser
import requests
from apscheduler.schedulers.background import BackgroundScheduler
from bs4 import BeautifulSoup
from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "top_stories.db"
MAX_CHECK_DOMAINS = 10
SESSION_PICKED_MEDIA_KEYS = "picked_media_keys"
SESSION_PICKED_DOMAINS = "picked_domains"
SETTINGS_LAST_SELECTION = "last_selected_domains"
SETTINGS_LAST_MEDIA_KEYS = "last_selected_media_keys"
SETTINGS_MEDIA_CATALOG_JSON = "media_catalog_json"
MEDIA_LIST_XLSX = BASE_DIR / "top story checker media list.xlsx"
# 模拟安卓端访问：必须用 feedparser 的 agent= 作为唯一 User-Agent。
# 若只传 request_headers["User-Agent"]，库仍会先加默认 feedparser UA，服务器往往只认第一条。
# 以下为常见「Android + Chrome Mobile」组合（系统版本与 Chrome 主版本对齐）。
MOBILE_HTTP_USER_AGENT = (
    "Mozilla/5.0 (Linux; Android 15; SM-S928B) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/133.0.0.0 Mobile Safari/537.36"
)
MOBILE_RSS_EXTRA_HEADERS = {
    "Accept-Language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7",
    "Sec-CH-UA": '"Google Chrome";v="133", "Chromium";v="133", "Not_A Brand";v="24"',
    "Sec-CH-UA-Mobile": "?1",
    "Sec-CH-UA-Platform": '"Android"',
    "Sec-CH-UA-Platform-Version": '"15.0.0"',
}
MOBILE_PAGE_REQUEST_HEADERS = {
    "User-Agent": MOBILE_HTTP_USER_AGENT,
    **MOBILE_RSS_EXTRA_HEADERS,
}
MAX_EXCEL_BYTES = 15 * 1024 * 1024
MAX_EXCEL_ROWS = 8000
DEFAULT_DOMAINS = [
    "cnn.com",
    "foxnews.com",
    "nytimes.com",
    "wsj.com",
    "washingtonpost.com",
    "latimes.com",
    "chicagotribune.com",
]

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-change-in-production")
scheduler = BackgroundScheduler(daemon=True)


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = get_conn()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS top_stories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fetched_at TEXT NOT NULL,
            domain TEXT NOT NULL,
            title TEXT NOT NULL,
            link TEXT NOT NULL,
            source TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_at TEXT NOT NULL,
            domains_count INTEGER NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )
    conn.commit()
    conn.close()
    ensure_top_stories_columns()


def ensure_top_stories_columns() -> None:
    conn = get_conn()
    cols = {row["name"] for row in conn.execute("PRAGMA table_info(top_stories)")}
    if "country" not in cols:
        conn.execute("ALTER TABLE top_stories ADD COLUMN country TEXT DEFAULT ''")
    if "media_name" not in cols:
        conn.execute("ALTER TABLE top_stories ADD COLUMN media_name TEXT DEFAULT ''")
    if "media_url" not in cols:
        conn.execute("ALTER TABLE top_stories ADD COLUMN media_url TEXT DEFAULT ''")
    conn.commit()
    conn.close()


def normalize_one_domain(raw: str) -> str:
    d = raw.strip().lower()
    if not d:
        return ""
    d = d.replace("https://", "").replace("http://", "").strip("/")
    return d


def normalize_domains(raw_text: str) -> list[str]:
    domains = []
    for line in raw_text.splitlines():
        d = normalize_one_domain(line)
        if d:
            domains.append(d)
    # Keep order, remove duplicates.
    return list(dict.fromkeys(domains))


def get_saved_domains() -> list[str]:
    conn = get_conn()
    row = conn.execute("SELECT value FROM settings WHERE key = 'domains'").fetchone()
    conn.close()
    if not row:
        return DEFAULT_DOMAINS
    domains = normalize_domains(row["value"])
    return domains or DEFAULT_DOMAINS


def save_domains(domains: list[str]) -> None:
    raw = "\n".join(domains)
    conn = get_conn()
    conn.execute(
        """
        INSERT INTO settings (key, value) VALUES ('domains', ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """,
        (raw,),
    )
    conn.commit()
    conn.close()


def merge_domain_lists(base: list[str], extra: list[str]) -> list[str]:
    """Append extra domains not already in base; preserve order."""
    seen = set(base)
    out = list(base)
    for d in extra:
        if d not in seen:
            seen.add(d)
            out.append(d)
    return out


def cell_to_domain(value: object) -> str:
    """Turn one spreadsheet cell into a hostname/domain, or empty string."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        return ""
    s = str(value).strip()
    if not s or s.lower() in ("none", "n/a", "-", "#n/a"):
        return ""
    if "://" in s:
        try:
            host = (urlparse(s).netloc or "").split("@")[-1].split(":")[0]
            return normalize_one_domain(host) if host else ""
        except Exception:
            return ""
    if "/" in s and "." in s:
        first = s.split("/")[0].strip()
        if "." in first and " " not in first:
            return normalize_one_domain(first)
    return normalize_one_domain(s)


def domains_from_excel_bytes(data: bytes) -> list[str]:
    """Read first worksheet; scan cells for URLs/domains (order preserved, deduped)."""
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        found: list[str] = []
        seen: set[str] = set()
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            row_count += 1
            if row_count > MAX_EXCEL_ROWS:
                break
            for cell in row:
                d = cell_to_domain(cell)
                if d and d not in seen:
                    seen.add(d)
                    found.append(d)
    finally:
        wb.close()
    return found


def normalize_catalog_url_key(raw: object) -> str:
    if raw is None:
        return ""
    return str(raw).strip()


def canonical_display_url(url_raw: str) -> str:
    u = url_raw.strip()
    if not u:
        return ""
    if "://" in u:
        return u
    if "." in u:
        return "https://" + u.lstrip("/")
    return u


def load_media_catalog_from_path(path: Path) -> list[dict[str, Any]]:
    """Read sheet 1: A 国家, B 媒体名, C 网址; 可选 D 手机版首页, E 头条 CSS 选择器（D+E 都有则优先抓手机站）。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    try:
        ws = wb.worksheets[0]
        for idx, row in enumerate(
            ws.iter_rows(min_row=1, max_col=5, values_only=True), start=1
        ):
            if idx > MAX_EXCEL_ROWS:
                break
            cells = (list(row) + [None, None, None, None, None])[:5]
            a, b, c, d, e = cells[0], cells[1], cells[2], cells[3], cells[4]
            country = "" if a is None else str(a).strip()
            name = "" if b is None else str(b).strip()
            url_raw = normalize_catalog_url_key(c)
            if not url_raw:
                continue
            # Skip typical header row
            if idx == 1:
                cl = url_raw.lower()
                if cl in ("url", "网址", "link", "website", "网址链接"):
                    continue
                if "country" in country.lower() and "media" in name.lower():
                    continue
            domain = cell_to_domain(url_raw)
            if not domain:
                continue
            key = url_raw
            if key in seen_keys:
                continue
            seen_keys.add(key)
            m_url_raw = normalize_catalog_url_key(d)
            headline_sel = "" if e is None else str(e).strip()
            m_url = canonical_display_url(m_url_raw) if m_url_raw else ""
            entry: dict[str, Any] = {
                "country": country,
                "name": name,
                "url": canonical_display_url(url_raw),
                "domain": domain,
                "key": key,
            }
            if m_url and headline_sel and urlparse(m_url).scheme in ("http", "https"):
                entry["m_url"] = m_url
                entry["headline_selector"] = headline_sel
            out.append(entry)
    finally:
        wb.close()
    return out


def save_media_catalog_json(catalog: list[dict[str, Any]]) -> None:
    raw = json.dumps(catalog, ensure_ascii=False)
    conn = get_conn()
    conn.execute(
        """
        INSERT INTO settings (key, value) VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """,
        (SETTINGS_MEDIA_CATALOG_JSON, raw),
    )
    conn.commit()
    conn.close()


def _media_catalog_from_settings() -> list[dict[str, Any]]:
    conn = get_conn()
    row = conn.execute(
        "SELECT value FROM settings WHERE key = ?", (SETTINGS_MEDIA_CATALOG_JSON,)
    ).fetchone()
    conn.close()
    if not row or not str(row["value"]).strip():
        return []
    try:
        data = json.loads(str(row["value"]))
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass
    return []


def get_media_catalog() -> list[dict[str, Any]]:
    """Prefer on-disk Excel next to the app; otherwise use last cached JSON in settings."""
    if MEDIA_LIST_XLSX.is_file():
        try:
            return load_media_catalog_from_path(MEDIA_LIST_XLSX)
        except Exception:
            return []
    return _media_catalog_from_settings()


def catalog_by_key(catalog: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {str(item["key"]): item for item in catalog if item.get("key")}


def get_last_selection_stored_keys() -> list[str]:
    conn = get_conn()
    row = conn.execute(
        "SELECT value FROM settings WHERE key = ?", (SETTINGS_LAST_MEDIA_KEYS,)
    ).fetchone()
    conn.close()
    if not row or not str(row["value"]).strip():
        return []
    keys = [normalize_catalog_url_key(k) for k in str(row["value"]).splitlines()]
    return [k for k in keys if k][:MAX_CHECK_DOMAINS]


def save_last_selection_keys(keys: list[str]) -> None:
    trimmed = keys[:MAX_CHECK_DOMAINS]
    raw = "\n".join(trimmed)
    conn = get_conn()
    conn.execute(
        """
        INSERT INTO settings (key, value) VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """,
        (SETTINGS_LAST_MEDIA_KEYS, raw),
    )
    conn.commit()
    conn.close()


def get_last_selection_media_keys_effective(catalog: list[dict[str, Any]]) -> list[str]:
    stored = get_last_selection_stored_keys()
    if not stored:
        return []
    allowed = {str(item["key"]) for item in catalog if item.get("key")}
    return [k for k in stored if k in allowed]


def format_selection_labels_for_keys(
    keys: list[str], catalog: list[dict[str, Any]]
) -> list[str]:
    by_k = catalog_by_key(catalog)
    labels: list[str] = []
    for k in keys:
        item = by_k.get(k)
        if not item:
            continue
        c = item.get("country") or "未填"
        n = item.get("name") or item.get("domain") or k
        labels.append(f"{c}：{n}")
    return labels


def get_last_selection_stored() -> list[str]:
    """Domains from the last successful manual run (max MAX_CHECK_DOMAINS)."""
    conn = get_conn()
    row = conn.execute(
        "SELECT value FROM settings WHERE key = ?", (SETTINGS_LAST_SELECTION,)
    ).fetchone()
    conn.close()
    if not row or not str(row["value"]).strip():
        return []
    return normalize_domains(str(row["value"]))[:MAX_CHECK_DOMAINS]


def save_last_selection(domains: list[str]) -> None:
    """Persist selection for scheduled runs and UI defaults."""
    trimmed = domains[:MAX_CHECK_DOMAINS]
    raw = "\n".join(trimmed)
    conn = get_conn()
    conn.execute(
        """
        INSERT INTO settings (key, value) VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """,
        (SETTINGS_LAST_SELECTION, raw),
    )
    conn.commit()
    conn.close()


def get_last_selection_effective() -> list[str]:
    """Last selection restricted to domains still present in the saved media list."""
    stored = get_last_selection_stored()
    if not stored:
        return []
    allowed = set(get_saved_domains())
    return [d for d in stored if d in allowed]


def google_news_rss_for_domain(domain: str) -> str:
    return (
        "https://news.google.com/rss/search?q="
        f"site:{domain}&hl=en-US&gl=US&ceid=US:en"
    )


def fetch_top_story(domain: str) -> dict:
    rss_url = google_news_rss_for_domain(domain)
    feed = feedparser.parse(
        rss_url,
        agent=MOBILE_HTTP_USER_AGENT,
        request_headers=MOBILE_RSS_EXTRA_HEADERS,
    )
    if feed.entries:
        top = feed.entries[0]
        return {
            "domain": domain,
            "title": top.get("title", "(No title)"),
            "link": top.get("link", ""),
            "source": "Google News RSS",
        }
    return {
        "domain": domain,
        "title": "No story found",
        "link": "",
        "source": "Google News RSS",
    }


def fetch_headline_from_mobile_homepage(
    news_site_m_url: str,
    headline_selector: str,
    domain: str,
) -> dict[str, Any] | None:
    """请求手机版首页，用 CSS 选择器定位头条区域，取区域内第一个 <a> 的文案与链接。"""
    try:
        parsed = urlparse(news_site_m_url)
        if parsed.scheme not in ("http", "https") or not parsed.netloc:
            return None
        resp = requests.get(
            news_site_m_url,
            headers=MOBILE_PAGE_REQUEST_HEADERS,
            timeout=15,
        )
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        nodes = soup.select(headline_selector)
        if not nodes:
            return None
        headline = nodes[0]
        a_tag = headline.find("a")
        if not a_tag:
            return None
        title = a_tag.get_text(strip=True)
        link = (a_tag.get("href") or "").strip()
        if not title or not link:
            return None
        if link.startswith("/"):
            link = urljoin(news_site_m_url, link)
        return {
            "domain": domain,
            "title": title,
            "link": link,
            "source": "媒体手机版官网头条",
        }
    except Exception:
        return None


def fetch_top_story_for_catalog_item(item: dict[str, Any]) -> dict[str, Any]:
    m_url = (item.get("m_url") or "").strip()
    sel = (item.get("headline_selector") or "").strip()
    if m_url and sel:
        scraped = fetch_headline_from_mobile_homepage(m_url, sel, item["domain"])
        if scraped:
            return {
                **scraped,
                "country": item.get("country") or "",
                "media_name": item.get("name") or "",
                "media_url": item.get("url") or "",
            }
    base = fetch_top_story(item["domain"])
    return {
        **base,
        "country": item.get("country") or "",
        "media_name": item.get("name") or "",
        "media_url": item.get("url") or "",
    }


def run_fetch_for_domains(domains: list[str]) -> None:
    rows: list[dict[str, Any]] = []
    for d in domains:
        r = fetch_top_story(d)
        r["country"] = ""
        r["media_name"] = ""
        r["media_url"] = ""
        rows.append(r)
    save_results(rows)
    export_latest_csv()


def run_fetch_for_catalog_items(items: list[dict[str, Any]]) -> None:
    rows = [fetch_top_story_for_catalog_item(item) for item in items]
    save_results(rows)
    export_latest_csv()


def save_results(rows: Iterable[dict]) -> None:
    now = dt.datetime.now().isoformat(timespec="seconds")
    conn = get_conn()
    count = 0
    for row in rows:
        conn.execute(
            """
            INSERT INTO top_stories (
                fetched_at, domain, title, link, source, country, media_name, media_url
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                now,
                row["domain"],
                row["title"],
                row["link"],
                row["source"],
                row.get("country", ""),
                row.get("media_name", ""),
                row.get("media_url", ""),
            ),
        )
        count += 1
    conn.execute("INSERT INTO runs (run_at, domains_count) VALUES (?, ?)", (now, count))
    conn.commit()
    conn.close()


def latest_rows() -> list[sqlite3.Row]:
    conn = get_conn()
    run = conn.execute("SELECT run_at FROM runs ORDER BY id DESC LIMIT 1").fetchone()
    if not run:
        conn.close()
        return []
    rows = conn.execute(
        """
        SELECT fetched_at, domain, country, media_name, media_url, title, link, source
        FROM top_stories
        WHERE fetched_at = ?
        ORDER BY domain
        """,
        (run["run_at"],),
    ).fetchall()
    conn.close()
    return rows


def export_latest_csv() -> Path | None:
    rows = latest_rows()
    if not rows:
        return None
    out_dir = BASE_DIR / "exports"
    out_dir.mkdir(exist_ok=True)
    stamp = rows[0]["fetched_at"].replace(":", "-")
    out_path = out_dir / f"top_stories_{stamp}.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "fetched_at",
                "country",
                "media_name",
                "media_url",
                "domain",
                "title",
                "link",
                "source",
            ]
        )
        for r in rows:
            writer.writerow(
                [
                    r["fetched_at"],
                    r["country"],
                    r["media_name"],
                    r["media_url"],
                    r["domain"],
                    r["title"],
                    r["link"],
                    r["source"],
                ]
            )
    return out_path


def _session_picked_media_keys() -> list[str]:
    raw = session.get(SESSION_PICKED_MEDIA_KEYS)
    if not isinstance(raw, list):
        return []
    out: list[str] = []
    for x in raw:
        k = normalize_catalog_url_key(x)
        if k and k not in out:
            out.append(k)
    return out[:MAX_CHECK_DOMAINS]


def _session_picked_domains() -> list[str]:
    raw = session.get(SESSION_PICKED_DOMAINS)
    if not isinstance(raw, list):
        return []
    out: list[str] = []
    for x in raw:
        d = normalize_one_domain(str(x))
        if d and d not in out:
            out.append(d)
    return out[:MAX_CHECK_DOMAINS]


def load_picked_display() -> tuple[list[dict[str, Any]], list[str], bool]:
    """从会话读取当前勾选，用于首页摘要与第二步页面。"""
    catalog = get_media_catalog()
    if catalog:
        keys = _session_picked_media_keys()
        if not keys:
            return [], [], False
        by_key = catalog_by_key(catalog)
        rows = [by_key[k] for k in keys if k in by_key]
        return rows, [], bool(rows)
    doms = _session_picked_domains()
    return [], doms, bool(doms)


@app.route("/cards")
def knowledge_cards():
    """Mobile-friendly zoology knowledge cards."""
    from flask import send_from_directory

    return send_from_directory(BASE_DIR / "static", "zoology-cards.html")


@app.route("/quiz")
def zoology_quiz():
    """Mobile-friendly advanced zoology multiple-choice cards."""
    from flask import send_from_directory

    return send_from_directory(BASE_DIR / "static", "zoology-quiz.html")


@app.route("/micro-ibo")
def micro_ibo_quiz():
    """Mobile-friendly IBO-style quiz based on the microscope image."""
    from flask import send_from_directory

    return send_from_directory(BASE_DIR / "static", "micro-ibo-quiz.html")


@app.route("/daily-IBO-test")
def daily_ibo_test():
    """Mobile-friendly daily IBO simulation test."""
    from flask import send_from_directory

    return send_from_directory(BASE_DIR / "static", "daily-IBO-test.html")


@app.route("/derivative-gaokao")
def derivative_gaokao_cards():
    """Mobile-friendly Gaokao derivative practice cards."""
    from flask import send_from_directory

    return send_from_directory(BASE_DIR / "static", "derivative-gaokao-cards.html")


@app.route("/wordcloud")
def wordcloud_page():
    """Chinese word cloud sized by trailing numeric weights."""
    from flask import send_from_directory

    return send_from_directory(BASE_DIR / "static", "wordcloud.html")


@app.route("/clock")
def day_night_clock():
    """Mobile-friendly 24h day/night circular clock with sun/moon pointer."""
    from flask import send_from_directory

    return send_from_directory(BASE_DIR / "static", "day-night-clock.html")


@app.route("/", methods=["GET"])
def home():
    rows = latest_rows()
    conn = get_conn()
    last_run = conn.execute("SELECT run_at FROM runs ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    schedule_time = f"{os.getenv('DAILY_RUN_HOUR', '08')}:{os.getenv('DAILY_RUN_MINUTE', '00')}"
    catalog = get_media_catalog()

    picked_rows, picked_domains, has_pick = load_picked_display()

    if catalog:
        keys_eff = get_last_selection_media_keys_effective(catalog)
        last_selection_labels = format_selection_labels_for_keys(keys_eff, catalog)
        if MEDIA_LIST_XLSX.is_file():
            catalog_hint = (
                f"已从工作目录读取《{MEDIA_LIST_XLSX.name}》，共 {len(catalog)} 条媒体。"
            )
        else:
            catalog_hint = (
                f"当前使用已缓存的媒体目录，共 {len(catalog)} 条。将《{MEDIA_LIST_XLSX.name}》"
                "放在程序同目录可改为读取本地文件。"
            )
    else:
        last_selection = get_last_selection_effective()
        last_selection_labels = last_selection
        catalog_hint = (
            f"未找到《{MEDIA_LIST_XLSX.name}》。请将该文件放在程序同目录，"
            "或在服务器上使用上次缓存的目录。"
        )

    return render_template(
        "index.html",
        media_catalog=catalog,
        picked_rows=picked_rows,
        picked_domains=picked_domains,
        has_pick=has_pick,
        max_check=MAX_CHECK_DOMAINS,
        last_selection_labels=last_selection_labels,
        rows=rows,
        last_run=last_run["run_at"] if last_run else None,
        schedule_time=schedule_time,
        catalog_hint=catalog_hint,
    )


@app.route("/step2", methods=["GET"])
def step_two():
    picked_rows, picked_domains, has_pick = load_picked_display()
    if not has_pick:
        flash("请先完成第一步：勾选媒体。", "error")
        return redirect(url_for("home"))
    return render_template(
        "step2.html",
        picked_rows=picked_rows,
        picked_domains=picked_domains,
        max_check=MAX_CHECK_DOMAINS,
    )


@app.route("/pick", methods=["GET", "POST"])
def pick_media():
    if request.method == "GET" and request.args.get("clear"):
        session.pop(SESSION_PICKED_MEDIA_KEYS, None)
        session.pop(SESSION_PICKED_DOMAINS, None)
        flash("已清空勾选，请重新选择媒体。", "success")
        return redirect(url_for("pick_media"))

    catalog = get_media_catalog()

    if request.method == "POST":
        if catalog:
            by_key = catalog_by_key(catalog)
            allowed = set(by_key.keys())
            keys: list[str] = []
            for s in request.form.getlist("selected_media_keys"):
                k = normalize_catalog_url_key(s)
                if k in allowed and k not in keys:
                    keys.append(k)
            if not keys:
                flash("请至少勾选一家媒体。", "error")
                return redirect(url_for("pick_media"))
            if len(keys) > MAX_CHECK_DOMAINS:
                flash(f"一次最多只能勾选 {MAX_CHECK_DOMAINS} 家媒体。", "error")
                return redirect(url_for("pick_media"))
            session[SESSION_PICKED_MEDIA_KEYS] = keys
            session.pop(SESSION_PICKED_DOMAINS, None)
        else:
            allowed = set(get_saved_domains())
            doms: list[str] = []
            for s in request.form.getlist("selected_domains"):
                d = normalize_one_domain(s)
                if d in allowed and d not in doms:
                    doms.append(d)
            if not doms:
                flash("请至少勾选一个域名。", "error")
                return redirect(url_for("pick_media"))
            if len(doms) > MAX_CHECK_DOMAINS:
                flash(f"一次最多只能勾选 {MAX_CHECK_DOMAINS} 个。", "error")
                return redirect(url_for("pick_media"))
            session[SESSION_PICKED_DOMAINS] = doms
            session.pop(SESSION_PICKED_MEDIA_KEYS, None)
        flash("第一步已完成。请打开「第二步」页面获取头条（地址 /step2）。", "success")
        return redirect(url_for("home"))

    # GET：展示完整列表供勾选
    last_keys = set(_session_picked_media_keys()) if catalog else set(_session_picked_domains())
    dl = get_saved_domains()
    return render_template(
        "pick.html",
        media_catalog=catalog,
        domain_list=dl if not catalog else [],
        max_check=MAX_CHECK_DOMAINS,
        last_selected_set=last_keys,
        checkbox_name="selected_media_keys" if catalog else "selected_domains",
    )


@app.route("/run", methods=["POST"])
def run_now():
    catalog = get_media_catalog()
    if catalog:
        keys = _session_picked_media_keys()
        if not keys:
            flash("请先在「第一步」勾选媒体页面完成选择，再在第二步页面（/step2）获取头条。", "error")
            return redirect(url_for("pick_media"))
        by_key = catalog_by_key(catalog)
        allowed = set(by_key.keys())
        keys = [k for k in keys if k in allowed]
        if not keys:
            flash("当前勾选已失效，请重新勾选媒体。", "error")
            return redirect(url_for("pick_media"))
        if len(keys) > MAX_CHECK_DOMAINS:
            flash(f"一次最多处理 {MAX_CHECK_DOMAINS} 家媒体，请返回第一步勾选页面减少数量。", "error")
            return redirect(url_for("pick_media"))
        save_last_selection_keys(keys)
        items = [by_key[k] for k in keys if k in by_key]
        run_fetch_for_catalog_items(items)
        flash("已根据勾选获取头条并更新结果。", "success")
        return redirect(url_for("home"))

    domains = _session_picked_domains()
    if not domains:
        flash("请先在「第一步」勾选媒体页面完成选择，再在第二步页面（/step2）获取头条。", "error")
        return redirect(url_for("pick_media"))
    allowed = set(get_saved_domains())
    domains = [d for d in domains if d in allowed]
    if not domains:
        flash("当前勾选已失效，请重新勾选。", "error")
        return redirect(url_for("pick_media"))
    if len(domains) > MAX_CHECK_DOMAINS:
        flash(f"一次最多处理 {MAX_CHECK_DOMAINS} 个域名，请返回第一步勾选页面减少数量。", "error")
        return redirect(url_for("pick_media"))
    save_last_selection(domains)
    run_fetch_for_domains(domains)
    flash("已根据勾选获取头条并更新结果。", "success")
    return redirect(url_for("home"))


@app.route("/import", methods=["POST"])
def import_domains():
    domains_text = request.form.get("domains", "")
    domains = normalize_domains(domains_text)
    if domains:
        save_domains(domains)
    return redirect(url_for("home"))


@app.route("/import-excel", methods=["POST"])
def import_excel():
    f = request.files.get("excel")
    if not f or not getattr(f, "filename", None):
        flash("请选择 Excel 文件（.xlsx 或 .xlsm）。", "error")
        return redirect(url_for("home"))
    name = f.filename.lower()
    if not (name.endswith(".xlsx") or name.endswith(".xlsm")):
        flash("仅支持 .xlsx / .xlsm 格式（不支持旧版 .xls）。", "error")
        return redirect(url_for("home"))
    try:
        data = f.read()
    except Exception as exc:
        flash(f"无法读取上传文件：{exc}", "error")
        return redirect(url_for("home"))
    if len(data) > MAX_EXCEL_BYTES:
        flash("Excel 文件过大（上限 15 MB）。", "error")
        return redirect(url_for("home"))
    try:
        from_excel = domains_from_excel_bytes(data)
    except Exception as exc:
        flash(f"无法解析 Excel：{exc}", "error")
        return redirect(url_for("home"))
    if not from_excel:
        flash("第一个工作表中没有识别到域名或网址。", "error")
        return redirect(url_for("home"))
    base = get_saved_domains()
    merged = merge_domain_lists(base, from_excel)
    added = len(merged) - len(base)
    save_domains(merged)
    flash(
        f"已从通用 Excel 合并：新增 {added} 个域名（文件中识别 {len(from_excel)} 个），"
        f"文本列表合计 {len(merged)} 个。",
        "success",
    )
    return redirect(url_for("home"))


@app.route("/export-domains", methods=["GET"])
def export_domains():
    domains = get_saved_domains()
    out_dir = BASE_DIR / "exports"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "domains_latest.txt"
    out_path.write_text("\n".join(domains), encoding="utf-8")
    return jsonify({"file": str(out_path), "count": len(domains)})


@app.route("/api/top-stories", methods=["GET"])
def top_stories_api():
    rows = latest_rows()
    data = [dict(r) for r in rows]
    return jsonify(data)


def start_scheduler() -> None:
    hour = int(os.getenv("DAILY_RUN_HOUR", "8"))
    minute = int(os.getenv("DAILY_RUN_MINUTE", "0"))

    def scheduled_job() -> None:
        catalog = get_media_catalog()
        if catalog:
            keys = get_last_selection_media_keys_effective(catalog)
            if not keys:
                return
            by_key = catalog_by_key(catalog)
            items = [by_key[k] for k in keys if k in by_key]
            if items:
                run_fetch_for_catalog_items(items)
        else:
            domains = get_last_selection_effective()
            if domains:
                run_fetch_for_domains(domains)

    if not scheduler.running:
        scheduler.add_job(
            scheduled_job,
            "cron",
            id="daily_top_stories_run",
            hour=hour,
            minute=minute,
            replace_existing=True,
        )
        scheduler.start()


def boot() -> None:
    init_db()
    save_domains(get_saved_domains())
    if MEDIA_LIST_XLSX.is_file():
        try:
            cat = load_media_catalog_from_path(MEDIA_LIST_XLSX)
            if cat:
                save_media_catalog_json(cat)
        except Exception:
            pass
    start_scheduler()


if __name__ == "__main__":
    boot()
    app.run(host="0.0.0.0", port=5000, debug=False)


if __name__ != "__main__":
    boot()
